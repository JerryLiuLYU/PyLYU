import csv
from openpyxl import load_workbook
def ReadCsv(src, result):
    #Open the csv file
    with open(src) as fin:
        #read the csv
        reader = csv.reader(fin)
        #enumerate the rows, so that you can
        #get the row index for the xlsx
        for index,row in enumerate(reader):
            #Assuming space separated,
            #Split the row to cells (column)
            #row = row[0].split(',')
            if len(row)>5:
                result[row[1]] = result.get(row[1],0)+ 1

def WriteResult(dest, result):
     #Open an xlsx for reading
    wb = load_workbook(filename = dest)
    #Get the current Active Sheet
    ws = wb.get_active_sheet()
    #You can also select a particular sheet
    #based on sheet name
    #ws = wb.get_sheet_by_name("Sheet1")
    for i in range(2,59): #从第2行到第58行，是学生的名单
        name = ws.cell(column=3, row=i).value
        count = result.get(name,0)
        ws.cell(column=4, row=i).value = count
        print(name)
    wb.save(dest)
    
result = {}#定义一个空字典，用来保存学生签订次数，例如{'张三':9,'李四':8,....}
for i in range(2,15):#循环读取本地的签到文件，2.csv---14.csv
    src = r'{0}.csv'.format(i) #构建文件名字
    ReadCsv(src, result) #将src指定的文件中学生签到数据，读取到result字典中
print(result) #打印出签到字典的结果，看看而已
WriteResult(r'summary.xlsx', result) #将result字典的签到次数，写入到最终的统计excel文件中